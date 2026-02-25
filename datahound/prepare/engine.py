from __future__ import annotations

import json
import re
import warnings
from dataclasses import asdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from datahound.download.types import DownloadConfig, PrepareConfig, PrepareTypeRule
import pyarrow as pa
import pyarrow.parquet as pq

# Suppress openpyxl warnings about missing default styles
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")


def read_master_columns(tables_dir: Path, master_filename: str) -> List[str]:
    path = tables_dir / master_filename
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return [str(c) if c is not None else "" for c in header]


def read_master_columns_any(tables_dir: Path, master_filename: str, prefer_parquet: bool = True) -> List[str]:
    # try parquet schema: companies/<company>/parquet/<Title>.parquet
    if prefer_parquet:
        parquet_path = tables_dir.parent / "parquet" / master_filename.replace(".xlsx", ".parquet")
        if parquet_path.exists():
            import pandas as pd
            df = pd.read_parquet(parquet_path)
            return [str(c) for c in df.columns.tolist()]
    return read_master_columns(tables_dir, master_filename)


def normalize_header(name: str) -> str:
    return str(name).strip().lower()


def _is_id_column(name: str) -> bool:
    normalized = normalize_header(name)
    if not normalized:
        return False
    if normalized in {"id", "primary id"}:
        return True
    return any(token in normalized for token in (" id", "_id", "#", "number"))


def extract_timestamp_from_name(name: str) -> Optional[datetime]:
    m = re.search(r"(\d{8}_\d{6}|\d{14})", name)
    if not m:
        return None
    ts = m.group(1)
    try:
        return datetime.strptime(ts, "%Y%m%d_%H%M%S") if "_" in ts else datetime.strptime(ts, "%Y%m%d%H%M%S")
    except Exception:
        return None


def pick_newest_by_type(files: List[Path], type_names: Dict[str, str]) -> Dict[str, Path]:
    newest: Dict[str, tuple[Path, datetime]] = {}
    for f in files:
        name_lower = f.name.lower()
        file_type: Optional[str] = None
        for needle, t in type_names.items():
            if needle in name_lower:
                file_type = t
                break
        if file_type is None:
            for t in set(type_names.values()):
                if t in name_lower:
                    file_type = t
                    break
        if not file_type:
            continue
        ts = extract_timestamp_from_name(f.name)
        if not ts:
            continue
        if file_type not in newest or ts > newest[file_type][1]:
            newest[file_type] = (f, ts)
    return {k: v[0] for k, v in newest.items()}


def load_source_file(path: Path) -> pd.DataFrame:
    """Load source file with data type preservation"""
    
    if path.suffix.lower() == ".csv":
        # Load CSV as strings to preserve formatting
        return pd.read_csv(path, dtype=str, na_filter=False)
    else:
        # Load Excel with special handling to preserve TRUE/FALSE as strings
        import openpyxl
        
        # Suppress openpyxl warnings locally
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
            
            # Load with openpyxl to preserve exact cell values
            wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        
        # Extract data preserving original string representations
        data = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # Header row
                headers = [str(cell) if cell is not None else "" for cell in row]
            else:
                # Data rows - convert everything to strings, preserving TRUE/FALSE
                row_data = []
                for cell in row:
                    if cell is None:
                        row_data.append("")
                    elif isinstance(cell, bool):
                        # Convert to master file format (Python boolean strings)
                        row_data.append("True" if cell else "False")
                    else:
                        row_data.append(str(cell))
                data.append(row_data)
        
        wb.close()
        
        # Create DataFrame
        if headers and data:
            df = pd.DataFrame(data, columns=headers)
        else:
            # Fallback to pandas if openpyxl fails
            df = pd.read_excel(path, dtype=str, na_filter=False)
            df = df.fillna("")
        
        return df


def apply_rules(df: pd.DataFrame, rules: PrepareTypeRule) -> pd.DataFrame:
    rename_map = {k: v for k, v in (rules.column_renames or {}).items()}
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def reorder_to_master(df: pd.DataFrame, master_columns: List[str], master_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    # map normalized headers to original
    norm_to_original: Dict[str, str] = {}
    for c in df.columns:
        norm_to_original[normalize_header(c)] = c
    ordered_cols = [norm_to_original.get(normalize_header(c), None) for c in master_columns]
    out = pd.DataFrame()
    
    # Get master file format examples for date columns
    master_date_formats = {}
    if master_df is not None:
        for col in master_columns:
            if 'date' in col.lower() and col in master_df.columns:
                # Get sample date format from master
                sample_dates = master_df[col].dropna().head(5)
                if len(sample_dates) > 0:
                    sample_date = str(sample_dates.iloc[0])
                    master_date_formats[col] = sample_date
    
    for mcol, ocol in zip(master_columns, ordered_cols):
        if ocol is None:
            out[mcol] = ""
        else:
            series = df[ocol].astype("string").fillna("")
            series = series.str.strip()
            if _is_id_column(mcol):
                series = series.str.replace(r'\.0$', '', regex=True)
            else:
                if 'date' in mcol.lower():
                    series = _format_dates_to_iso(series)
                series = series.str.replace(r'^(\d+)\.0$', r'\1', regex=True)
            out[mcol] = series.astype("string")
    return out


def _format_dates_to_iso(series: pd.Series) -> pd.Series:
    """Format dates to ISO datetime format (YYYY-MM-DD HH:MM:SS)"""
    
    try:
        formatted_series = []
        
        for value in series:
            value_str = str(value).strip()
            
            if not value_str or value_str in ['', 'nan', 'NaT', 'None', 'NULL']:
                formatted_series.append("")
                continue
            
            try:
                # Parse the date (handles various input formats)
                import warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", message="Could not infer format")
                    parsed_date = pd.to_datetime(value_str, errors='coerce')
                
                if pd.notna(parsed_date):
                    # Format to ISO datetime string
                    formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                    formatted_series.append(formatted_date)
                else:
                    # Keep original if can't parse
                    formatted_series.append(value_str)
                    
            except Exception:
                # Keep original if any error
                formatted_series.append(value_str)
        
        return pd.Series(formatted_series)
        
    except Exception:
        # Return original series if any error
        return series


def drop_trailing_row_if_needed(df: pd.DataFrame, drop: bool) -> pd.DataFrame:
    if not drop or df.empty:
        return df
    return df.iloc[:-1, :] if len(df) > 1 else df.iloc[0:0]


def prepare_latest_files(config: DownloadConfig, selected_types: Optional[List[str]] = None, write_csv: bool = False, write_parquet: bool = True) -> Dict[str, Path]:
    assert config.prepare is not None
    prep: PrepareConfig = config.prepare
    data_dir = Path(config.data_dir)
    out_dir = data_dir  # prepared files stay in same dir prefixed with prepared_

    all_files = list(data_dir.glob("*.xlsx")) + list(data_dir.glob("*.csv"))
    type_names = {
        "job_events": "jobs",
        "customer_events": "customers",
        "membership_events": "memberships",
        "call_events": "calls",
        "estimate_events": "estimates",
        "invoice_events": "invoices",
        "location_events": "locations",
    }
    newest = pick_newest_by_type(all_files, type_names)
    if selected_types:
        newest = {t: p for t, p in newest.items() if t in selected_types}

    from central_logging.config import pipeline_dir
    results: Dict[str, Path] = {}
    log_path = pipeline_dir(config.company) / "prepare.jsonl"
    for ftype, fpath in newest.items():
        try:
            if not (write_csv or write_parquet):
                _append_log(log_path, {
                    "ts": datetime.now(UTC).isoformat(),
                    "company": config.company,
                    "file_type": ftype,
                    "status": "skipped_no_output_format_selected",
                    "source": fpath.name,
                })
                continue
            
            # Check if output file(s) already exist and are up-to-date
            source_mtime = fpath.stat().st_mtime
            should_skip = False
            existing_output = None
            
            if write_parquet:
                expected_parquet = out_dir / f"prepared_{fpath.stem}.parquet"
                if expected_parquet.exists():
                    parquet_mtime = expected_parquet.stat().st_mtime
                    # Skip if parquet is newer than or equal to source file
                    if parquet_mtime >= source_mtime:
                        should_skip = True
                        existing_output = expected_parquet
            
            # Also check CSV if parquet check didn't skip and CSV is requested
            if not should_skip and write_csv:
                expected_csv = out_dir / f"prepared_{fpath.stem}.csv"
                if expected_csv.exists():
                    csv_mtime = expected_csv.stat().st_mtime
                    # Skip if CSV is newer than or equal to source file
                    if csv_mtime >= source_mtime:
                        should_skip = True
                        existing_output = expected_csv
            
            if should_skip and existing_output:
                _append_log(log_path, {
                    "ts": datetime.now(UTC).isoformat(),
                    "company": config.company,
                    "file_type": ftype,
                    "status": "skipped_already_prepared",
                    "source": fpath.name,
                    "existing_output": existing_output.name,
                    "source_mtime": datetime.fromtimestamp(source_mtime, UTC).isoformat(),
                    "output_mtime": datetime.fromtimestamp(existing_output.stat().st_mtime, UTC).isoformat(),
                })
                results[ftype] = existing_output
                continue
            
            master_name = prep.file_type_to_master.get(ftype)
            if not master_name:
                _append_log(log_path, {
                    "ts": datetime.now(UTC).isoformat(),
                    "company": config.company,
                    "file_type": ftype,
                    "status": "skipped_no_master",
                    "source": fpath.name,
                })
                continue
            master_cols = read_master_columns_any(prep.tables_dir, master_name, prefer_parquet=True)
            
            # Load master DataFrame to get format examples
            master_df = None
            try:
                # Try to load master parquet file for format reference
                parquet_dir = Path(prep.tables_dir).parent / "parquet"
                master_parquet = parquet_dir / master_name.replace(".xlsx", ".parquet")
                if master_parquet.exists():
                    master_df = pd.read_parquet(master_parquet)
            except Exception:
                master_df = None
            
            df = load_source_file(fpath)
            rules = prep.type_rules.get(ftype, PrepareTypeRule())
            df = apply_rules(df, rules)
            # Convert configured date columns if present with robust parsing
            if getattr(rules, "date_columns", None):
                for dc in rules.date_columns:
                    if dc in df.columns:
                        series = df[dc]
                        # normalize placeholders (treat common tokens as empty)
                        extra_ph = ["NaT", "nan", "None"]
                        for ph in (getattr(rules, "date_placeholders", []) or []) + extra_ph:
                            series = series.replace(ph, "", regex=False)
                        parsed = None
                        # try configured formats; accept the first that yields any valid values
                        for fmt in (getattr(rules, "date_formats", []) or []):
                            try:
                                candidate = pd.to_datetime(series, format=fmt, errors="coerce")
                            except Exception:
                                candidate = None
                            if candidate is not None and candidate.notna().any():
                                parsed = candidate
                                break
                        # try generic parser
                        if parsed is None or not parsed.notna().any():
                            try:
                                candidate = pd.to_datetime(series, errors="coerce")
                                if candidate.notna().any():
                                    parsed = candidate
                            except Exception:
                                parsed = None
                        # excel serial fallback
                        if (parsed is None or not parsed.notna().any()) and getattr(rules, "excel_serial", False):
                            try:
                                numeric = pd.to_numeric(series, errors="coerce")
                                # Excel serial origin 1899-12-30
                                origin = pd.Timestamp("1899-12-30")
                                candidate = origin + pd.to_timedelta(numeric, unit="D")
                                if candidate.notna().any():
                                    parsed = candidate
                            except Exception:
                                parsed = None
                        if parsed is not None:
                            # Keep original format if it was already in a valid format
                            # Only apply parsing if the original values were clearly invalid
                            original_series = df[dc].astype(str).fillna("")
                            
                            # Check if original values look like valid dates
                            original_looks_valid = original_series.str.match(r'\d{1,2}/\d{1,2}/\d{4}').any()
                            
                            if original_looks_valid:
                                # Keep original format to avoid formatting changes
                                continue
                            else:
                                # Apply parsed dates only if originals were clearly invalid
                                df[dc] = parsed
            # Fallback fix for calls: map 'ID' -> 'Call ID' if not already present
            if ftype == "calls" and ("Call ID" not in df.columns) and ("ID" in df.columns):
                df["Call ID"] = df["ID"]
            df = drop_trailing_row_if_needed(df, rules.drop_last_row)
            out_df = reorder_to_master(df, master_cols, master_df)
            csv_path = None
            parquet_out = None
            if write_csv:
                out_name = f"prepared_{fpath.stem}.csv"
                out_path = out_dir / out_name
                out_df.to_csv(out_path, index=False)
                csv_path = out_path
            if write_parquet:
                try:
                    parquet_out = out_dir / (f"prepared_{fpath.stem}.parquet")
                    pq.write_table(pa.Table.from_pandas(out_df), parquet_out)
                except Exception:
                    parquet_out = None
            # pick primary result: parquet first
            results[ftype] = parquet_out or csv_path or out_dir / f"prepared_{fpath.stem}.csv"
            _append_log(log_path, {
                "ts": datetime.now(UTC).isoformat(),
                "company": config.company,
                "file_type": ftype,
                "status": "prepared",
                "source": fpath.name,
                "outputs": {
                    "csv": (csv_path.name if csv_path else None),
                    "parquet": (parquet_out.name if parquet_out else None),
                },
            })
        except Exception as e:
            _append_log(log_path, {
                "ts": datetime.now(UTC).isoformat(),
                "company": config.company,
                "file_type": ftype,
                "status": "error",
                "source": fpath.name,
                "error": str(e),
            })
    return results


def _append_log(path: Path, record: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record) + "\n")


