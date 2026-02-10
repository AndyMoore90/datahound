from __future__ import annotations

import json
from datetime import UTC, datetime
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import shutil

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from datahound.upsert.types import AuditChange, UpsertResult


def _is_id_like(column_name: str, id_col: str) -> bool:
    name = (column_name or "").lower()
    if column_name == id_col:
        return True
    return ("id" in name) or ("#" in column_name) or ("number" in name)


def _singularize_type_name(ftype: str) -> str:
    mapping = {
        "jobs": "job",
        "customers": "customer",
        "memberships": "membership",
        "calls": "call",
        "estimates": "estimate",
        "invoices": "invoice",
        "locations": "location",
    }
    return mapping.get(ftype, ftype[:-1] if ftype.endswith("s") else ftype)


def _canonicalize_series_for_compare(series: pd.Series, column_name: str, id_col: str) -> pd.Series:
    s = series.astype("string").fillna("").str.strip()
    # map booleans to numeric-like for comparison to avoid false diffs
    s_lower = s.str.lower()
    # normalize placeholder empties (avoid "NaT"/"nan" showing as updates against empty)
    empty_tokens = {"", "nat", "nan", "none", "null", "n/a", "-"}
    s = s.where(~s_lower.isin(empty_tokens), "")
    s_lower = s.str.lower()
    s = s.where(~s_lower.isin(["true", "false"]), s_lower.map({"true": "1", "false": "0"}))
    # strip float tails like 123.0 → 123 for comparison
    s = s.str.replace(r"^(\d+)\.0+$", r"\1", regex=True)
    # normalize dates to ISO for comparison when parsable
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Could not infer format")
            dt = pd.to_datetime(s, errors="coerce")
            iso = dt.dt.strftime("%Y-%m-%d")
            s = s.where(dt.isna(), iso)
    except Exception:
        pass
    # ID-like extra normalization
    if _is_id_like(column_name, id_col):
        s = s.str.replace(r"\.0$", "", regex=True)
    return s


def _detect_date_pattern(s: str) -> Optional[str]:
    if isinstance(s, str):
        if pd.notna(s) and len(s) >= 8:
            if pd.Series([s]).str.match(r"^\d{4}-\d{2}-\d{2}$").iloc[0]:
                return "iso"
            if pd.Series([s]).str.match(r"^\d{1,2}/\d{1,2}/\d{4}$").iloc[0]:
                return "mdy_slash"
    return None


def _infer_master_formats(master_df: pd.DataFrame, columns: List[str]) -> Dict[str, Dict[str, str]]:
    formats: Dict[str, Dict[str, str]] = {}
    for col in columns:
        sample = None
        if col in master_df.columns:
            series = master_df[col].astype("string").dropna().astype(str).str.strip()
            if not series.empty:
                sample = series.iloc[0]
        kind = None
        fmt = None
        if sample is not None and sample != "":
            low = str(sample).lower()
            if low in ("true", "false"):
                kind = "bool"; fmt = "tf"
            elif low in ("yes", "no"):
                kind = "bool"; fmt = "yn"
            elif low in ("1", "0"):
                kind = "bool"; fmt = "10"
            else:
                dpat = _detect_date_pattern(str(sample))
                if dpat is not None:
                    kind = "date"; fmt = dpat
                elif pd.Series([str(sample)]).str.match(r"^\d+$").iloc[0]:
                    kind = "int"; fmt = "digits"
        if kind is not None:
            formats[col] = {"kind": kind, "fmt": fmt}
    return formats


def _format_to_master(value: object, fmtinfo: Optional[Dict[str, str]]) -> str:
    s = "" if pd.isna(value) else str(value)
    if not fmtinfo:
        return s
    kind = fmtinfo.get("kind")
    fmt = fmtinfo.get("fmt")
    if kind == "bool":
        low = s.lower()
        truthy = {"true", "1", "yes", "y"}
        falsy = {"false", "0", "no", "n"}
        if fmt == "tf":
            if low in truthy:
                return "True"
            if low in falsy:
                return "False"
        if fmt == "yn":
            if low in truthy:
                return "Yes"
            if low in falsy:
                return "No"
        if fmt == "10":
            if low in truthy:
                return "1"
            if low in falsy:
                return "0"
        return s
    if kind == "date":
        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.isna(dt):
                return s
            if fmt == "iso":
                return dt.strftime("%Y-%m-%d")
            if fmt == "mdy_slash":
                return dt.strftime("%m/%d/%Y")
            return s
        except Exception:
            return s
    if kind == "int":
        # strip trailing .0
        return pd.Series([s]).str.replace(r"^(\d+)\.0+$", r"\1", regex=True).iloc[0]
    return s


def read_master_excel(master_path: Path) -> Tuple[List[str], pd.DataFrame]:
    wb = load_workbook(master_path, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    wb.close()
    df = pd.DataFrame(rows, columns=header)
    return [str(c) if c is not None else "" for c in header], df


def _type_title(ftype: str, master_filename: Optional[str]) -> str:
    if master_filename:
        return master_filename.replace(".xlsx", "")
    return ftype.capitalize()


def _parquet_path_for_type(tables_dir: Path, ftype: str, master_filename: Optional[str]) -> Path:
    # companies/<company>/tables → companies/<company>/parquet
    parquet_dir = tables_dir.parent / "parquet"
    parquet_dir.mkdir(parents=True, exist_ok=True)
    name = _type_title(ftype, master_filename)
    return parquet_dir / f"{name}.parquet"


def read_master_any(tables_dir: Path, ftype: str, master_filename: Optional[str], prefer_parquet: bool) -> Tuple[str, List[str], pd.DataFrame, Path]:
    """Return storage kind (parquet|excel), header, df, path."""
    parquet_path = _parquet_path_for_type(tables_dir, ftype, master_filename)
    if prefer_parquet and parquet_path.exists():
        df = pd.read_parquet(parquet_path)
        header = [str(c) for c in df.columns.tolist()]
        return "parquet", header, df, parquet_path
    master_path = tables_dir / (master_filename if master_filename else f"{ftype.capitalize()}.xlsx")
    if master_path.exists():
        header, df = read_master_excel(master_path)
        return "excel", header, df, master_path
    # fallback: if parquet exists but prefer_parquet False
    if parquet_path.exists():
        df = pd.read_parquet(parquet_path)
        header = [str(c) for c in df.columns.tolist()]
        return "parquet", header, df, parquet_path
    return "missing", [], pd.DataFrame(), parquet_path


def write_master_excel(master_path: Path, header: List[str], rows: List[List[object]], backup: bool = True, backup_dir: Optional[Path] = None) -> None:
    from openpyxl import Workbook
    tmp_path = master_path.with_suffix(".tmp.xlsx")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(header)
    # sanitize values to avoid <NA> / unsupported types in Excel
    sanitized_rows: List[List[object]] = []
    for r in rows:
        sanitized_rows.append([("" if pd.isna(v) else str(v)) for v in r])
    for r in sanitized_rows:
        ws.append(r)
    wb.save(tmp_path)
    wb.close()
    if backup and master_path.exists():
        did_external = False
        if backup_dir is not None:
            try:
                backup_dir.mkdir(parents=True, exist_ok=True)
                external_backup = backup_dir / f"{master_path.stem}_backup_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.xlsx"
                shutil.copy2(master_path, external_backup)
                did_external = True
            except Exception:
                did_external = False
        if not did_external:
            backup_path = master_path.with_name(f"{master_path.stem}_backup_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.xlsx")
            try:
                master_path.replace(backup_path)
            except Exception:
                pass
    tmp_path.replace(master_path)


def write_master_parquet(parquet_path: Path, df: pd.DataFrame, backup: bool = True, backup_dir: Optional[Path] = None) -> None:
    if backup and parquet_path.exists():
        try:
            bak_dir = backup_dir if backup_dir is not None else parquet_path.parent
            bak_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.now(UTC).strftime('%Y%m%d_%H%M%S')
            shutil.copy2(parquet_path, bak_dir / f"{parquet_path.stem}_backup_{ts}.parquet")
        except Exception:
            pass
    # ensure all values are strings to keep behavior consistent
    df2 = df.copy()
    for c in df2.columns:
        df2[c] = df2[c].astype("string").fillna("")
    pa_table = pa.Table.from_pandas(df2, preserve_index=False)
    pq.write_table(pa_table, parquet_path)


def _update_master_inplace(master_path: Path, id_col: str, master_header: List[str], master_df: pd.DataFrame, prepared_df: pd.DataFrame) -> None:
    # openpyxl read/write mode: update existing rows by ID, append new IDs at end
    from openpyxl import load_workbook
    wb = load_workbook(master_path)
    ws = wb.active
    # build column indices map
    col_to_idx = {col: idx + 1 for idx, col in enumerate(master_header)}
    id_idx = col_to_idx.get(id_col)
    if id_idx is None:
        wb.close()
        raise RuntimeError(f"ID column not found in master header: {id_col}")
    # map ID -> row number for master (starting at row 2)
    id_to_row: Dict[str, int] = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        id_val = row[id_idx - 1]
        if id_val is None:
            continue
        id_to_row[str(id_val)] = r_idx
    # update existing
    for _id, p_row in prepared_df.set_index(id_col).iterrows():
        _sid = str(_id)
        r_idx = id_to_row.get(_sid)
        if r_idx is None:
            continue
        for col in master_header:
            c_idx = col_to_idx[col]
            ws.cell(row=r_idx, column=c_idx, value=str(p_row[col]) if col in p_row else "")
    # append new ids
    last_row = ws.max_row
    for _id, p_row in prepared_df.set_index(id_col).iterrows():
        _sid = str(_id)
        if _sid in id_to_row:
            continue
        last_row += 1
        # write entire row according to header order
        for col in master_header:
            c_idx = col_to_idx[col]
            ws.cell(row=last_row, column=c_idx, value=str(p_row[col]) if col in p_row else "")
    wb.save(master_path)
    wb.close()


def load_prepared(path: Path) -> pd.DataFrame:
    suf = path.suffix.lower()
    if suf == ".csv":
        return pd.read_csv(path)
    if suf in (".xlsx", ".xls"):
        return pd.read_excel(path)
    if suf == ".parquet":
        return pd.read_parquet(path)
    # fallback attempt
    return pd.read_csv(path)


def compute_audit_changes(merged: pd.DataFrame, id_col: str, master_cols: List[str]) -> List[AuditChange]:
    changes: List[AuditChange] = []
    for col in master_cols:
        mcol = f"{col}_x"
        pcol = f"{col}_y"
        if mcol in merged.columns and pcol in merged.columns:
            # normalize to comparable strings, with ID-like float tail stripping
            lhs = _canonicalize_series_for_compare(merged[mcol], col, id_col)
            rhs = _canonicalize_series_for_compare(merged[pcol], col, id_col)
            diff = (lhs != rhs)
            for i, is_different in enumerate(diff):
                    if is_different:
                        row = merged.iloc[i]
                        oldv = row[mcol]
                        newv = row[pcol]
                        changes.append(AuditChange(id_value=str(row[id_col]), column=col, old_value="" if pd.isna(oldv) else str(oldv), new_value="" if pd.isna(newv) else str(newv)))
    return changes


def build_updated_master(master_df: pd.DataFrame, prepared_df: pd.DataFrame, id_col: str, master_cols: List[str]) -> pd.DataFrame:
    master_df = master_df.copy()
    prepared_df = prepared_df.copy()
    # normalize all columns to string for stable compare/write
    for col in set(master_cols + [id_col]):
        if col in master_df.columns:
            master_df[col] = master_df[col].fillna("").astype(str).str.strip()
        if col in prepared_df.columns:
            prepared_df[col] = prepared_df[col].fillna("").astype(str).str.strip()
    master_df.set_index(id_col, inplace=True, drop=False)
    prepared_df.set_index(id_col, inplace=True, drop=False)
    # replace existing IDs with prepared rows where present
    common_ids = master_df.index.intersection(prepared_df.index)
    master_df.loc[common_ids, master_cols] = prepared_df.loc[common_ids, master_cols]
    # append new ids
    new_ids = prepared_df.index.difference(master_df.index)
    # ensure we don't duplicate id column in selection
    if id_col in master_cols:
        select_cols = master_cols
    else:
        select_cols = master_cols + [id_col]
    appended = pd.concat([master_df, prepared_df.loc[new_ids, select_cols]], axis=0)
    # preserve order: existing first (original order), then new ids in prepared order
    appended = appended.drop_duplicates(subset=[id_col], keep="last")
    return appended[master_cols]


def apply_cellwise_updates(master_df: pd.DataFrame, prepared_df: pd.DataFrame, id_col: str, master_cols: List[str], changes: List[AuditChange]) -> pd.DataFrame:
    updated = master_df.copy()
    if id_col not in updated.columns:
        return updated
    updated.set_index(id_col, inplace=True, drop=False)
    p_indexed = prepared_df.set_index(id_col, drop=False)
    for ch in changes:
        if ch.id_value in p_indexed.index and ch.column in master_cols and ch.column in p_indexed.columns:
            try:
                updated.at[ch.id_value, ch.column] = "" if pd.isna(p_indexed.at[ch.id_value, ch.column]) else str(p_indexed.at[ch.id_value, ch.column])
            except Exception:
                continue
    # append new ids
    new_ids = p_indexed.index.difference(updated.index)
    if len(new_ids) > 0:
        cols_for_append = [c for c in master_cols if c in p_indexed.columns]
        to_append = p_indexed.loc[new_ids, cols_for_append].copy()
        for c in master_cols:
            if c in to_append.columns:
                to_append[c] = to_append[c].astype("string").fillna("")
        appended = pd.concat([updated, to_append], axis=0)
    else:
        appended = updated
    appended.reset_index(drop=True, inplace=True)
    return appended[master_cols]


def upsert_type(company: str, data_dir: Path, tables_dir: Path, ftype: str, prepared_path: Path, id_col: str, event_rules: Dict[str, Dict] | None = None, master_filename: Optional[str] = None, backup: bool = True, backup_dir: Optional[Path] = None, write_mode: str = "inplace", dry_run: bool = False, maintain_store: bool = False, prefer_parquet: bool = True, excel_export: bool = False) -> UpsertResult:
    storage, master_header, master_df_initial, master_path_obj = read_master_any(tables_dir, ftype, master_filename, prefer_parquet)
    # log start
    logs_dir = Path(data_dir) / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_path = logs_dir / "apply_log.jsonl"
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(json.dumps({
            "ts": datetime.now(UTC).isoformat(),
            "company": company,
            "file_type": ftype,
            "status": "start",
            "storage": storage,
            "master": str(master_path_obj),
            "prepared": prepared_path.name,
        }) + "\n")
    if storage == "missing":
        # log skip: master missing
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps({
                "ts": datetime.now(UTC).isoformat(),
                "company": company,
                "file_type": ftype,
                "status": "skipped_master_not_found",
                "master": str(master_path_obj),
                "prepared": prepared_path.name,
            }) + "\n")
        return UpsertResult(type_name=ftype, examined_rows=0, updated_rows=0, new_rows=0, audit_changes=[], events_emitted=0, output_master_path=None)
    t0 = time.perf_counter()
    master_header, master_df = master_header, master_df_initial
    t1 = time.perf_counter()
    prepared_df = load_prepared(prepared_path)
    t2 = time.perf_counter()
    # align to master columns (missing columns become empty)
    for col in master_header:
        if col not in prepared_df.columns:
            prepared_df[col] = ""
    prepared_df = prepared_df[master_header]
    # ensure key types align for merge
    if id_col in master_df.columns:
        master_df[id_col] = master_df[id_col].astype("string")
        master_df[id_col] = _canonicalize_series_for_compare(master_df[id_col], id_col, id_col)
    if id_col in prepared_df.columns:
        prepared_df[id_col] = prepared_df[id_col].astype("string")
        prepared_df[id_col] = _canonicalize_series_for_compare(prepared_df[id_col], id_col, id_col)
    # merge for change detection
    merged = master_df.merge(prepared_df, how="outer", on=id_col, indicator=True, suffixes=("_x", "_y"))
    # metrics
    examined = len(master_df)
    new_rows = merged[merged["_merge"] == "right_only"][id_col].dropna().astype(str).nunique()
    audit_changes = compute_audit_changes(merged[merged["_merge"] == "both"], id_col, master_header)
    # decoupled: do not emit events here; log granular changes instead
    events_count = 0
    # build updated master
    # build updated master using cell-level changes, formatting to match existing master patterns
    fmt = _infer_master_formats(master_df, master_header)
    # conform prepared values to master formats column-wise
    prepared_conformed = prepared_df.copy()
    for c in master_header:
        if c in prepared_conformed.columns:
            prepared_conformed[c] = prepared_conformed[c].map(lambda v: _format_to_master(v, fmt.get(c)))
    updated_master = apply_cellwise_updates(master_df, prepared_conformed, id_col, master_header, audit_changes)
    updated_count = len(audit_changes)
    t3 = time.perf_counter()
    if not dry_run:
        if storage == "parquet":
            write_master_parquet(master_path_obj, updated_master, backup=backup, backup_dir=backup_dir)
            if excel_export:
                # optional Excel export next to tables dir
                excel_path = tables_dir / f"{_type_title(ftype, master_filename)}.xlsx"
                write_master_excel(excel_path, master_header, updated_master.values.tolist(), backup=backup, backup_dir=backup_dir)
            used_inplace = False
        else:
            if write_mode == "inplace":
                if backup:
                    write_master_excel(master_path_obj, master_header, master_df[master_header].values.tolist(), backup=True, backup_dir=backup_dir)
                _update_master_inplace(master_path_obj, id_col, master_header, master_df, prepared_df)
                used_inplace = True
            else:
                write_master_excel(master_path_obj, master_header, updated_master.values.tolist(), backup=backup, backup_dir=backup_dir)
                used_inplace = False
    else:
        used_inplace = (write_mode == "inplace")
    t4 = time.perf_counter()
    # determine final output path that was written/targeted
    output_path = master_path_obj
    # logging
    logs_dir = Path(data_dir) / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_path = logs_dir / "apply_log.jsonl"
    # per-type changes file
    singular = _singularize_type_name(ftype)
    changes_log_path = logs_dir / f"{singular}_changes_log.jsonl"
    # per-change logs for updates
    with open(changes_log_path, "a", encoding="utf-8") as cf:
        for ch in audit_changes:
            cf.write(json.dumps({
                "ts": datetime.now(UTC).isoformat(),
                "company": company,
                "file_type": ftype,
                "change_type": "update_cell",
                "id": ch.id_value,
                "column": ch.column,
                "old": ch.old_value,
                "new": ch.new_value,
                "master": str(output_path),
                "prepared": prepared_path.name,
            }) + "\n")
    # per-row logs for inserts
    new_ids_series = merged[merged["_merge"] == "right_only"][id_col].dropna().astype(str)
    if not new_ids_series.empty:
        new_ids_list = new_ids_series.tolist()
        for nid in new_ids_list:
            try:
                row_series = prepared_df.set_index(id_col).loc[nid]
                row_dict = {col: ("" if pd.isna(row_series[col]) else str(row_series[col])) for col in master_header}
            except Exception:
                row_dict = {}
            with open(changes_log_path, "a", encoding="utf-8") as cf:
                cf.write(json.dumps({
                    "ts": datetime.now(UTC).isoformat(),
                    "company": company,
                    "file_type": ftype,
                    "change_type": "insert_row",
                    "id": nid,
                    "row": row_dict,
                    "master": str(output_path),
                    "prepared": prepared_path.name,
                }) + "\n")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(json.dumps({
            "ts": datetime.now(UTC).isoformat(),
            "company": company,
            "file_type": ftype,
            "examined_rows": examined,
            "new_rows": int(new_rows),
            "audit_changes": len(audit_changes),
            "events": events_count,
            "master": str(output_path),
            "prepared": prepared_path.name,
            "backup": bool(backup),
            "backup_dir": str(backup_dir) if backup_dir else None,
            "write_mode": write_mode,
            "dry_run": bool(dry_run),
            "ms_read_master": int((t1 - t0) * 1000),
            "ms_read_prepared": int((t2 - t1) * 1000),
            "ms_diff": int((t3 - t2) * 1000),
            "ms_write": int((t4 - t3) * 1000),
        }) + "\n")
    return UpsertResult(
        type_name=ftype,
        examined_rows=examined,
        updated_rows=updated_count,
        new_rows=int(new_rows),
        audit_changes=audit_changes,
        events_emitted=events_count,
        output_master_path=output_path if not dry_run else None,
        ms_read_master=int((t1 - t0) * 1000),
        ms_read_prepared=int((t2 - t1) * 1000),
        ms_diff=int((t3 - t2) * 1000),
        ms_write=int((t4 - t3) * 1000),
        used_inplace_write=used_inplace,
        dry_run=bool(dry_run),
    )


def find_latest_prepared(data_dir: Path, ftype: str, prefer_parquet: bool = False) -> Optional[Path]:
    seeds_by_type: Dict[str, List[str]] = {
        "jobs": ["jobs", "job_events"],
        "customers": ["customers", "customer_events"],
        "memberships": ["memberships", "membership_events"],
        "calls": ["calls", "call_events"],
        "estimates": ["estimates", "estimate_events"],
        "invoices": ["invoices", "invoice_events"],
        "locations": ["locations", "location_events"],
    }
    seeds = [s.lower() for s in seeds_by_type.get(ftype, [ftype])]  # fall back to canonical only
    ext_order = [".parquet", ".csv", ".xlsx"] if prefer_parquet else [".csv", ".xlsx", ".parquet"]
    for ext in ext_order:
        files = list(Path(data_dir).glob(f"prepared_*{ext}"))
        matched = []
        for p in files:
            name = p.name.lower()
            if any(seed in name for seed in seeds):
                matched.append(p)
        if matched:
            matched.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            return matched[0]
    return None


